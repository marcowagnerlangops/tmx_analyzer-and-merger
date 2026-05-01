from __future__ import annotations

import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET
import tempfile
import unicodedata
from io import BytesIO

import streamlit as st


try:
    import pandas as pd
except Exception as e:
    raise SystemExit("pandas is required. Install with: pip install pandas") from e

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except Exception as e:
    raise SystemExit("openpyxl and pillow are required. Install with: pip install openpyxl pillow") from e

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except Exception as e:
    raise SystemExit("matplotlib is required. Install with: pip install matplotlib") from e


APP_TITLE = "TMX Analyzer – Phase 2 Advanced"
APP_GEOMETRY = "1820x1060"
MAKER_LINE = "Made by LangOps Solutions"
PAGE_SIZE_DEFAULT = 200

DUP_YELLOW = PatternFill(fill_type="solid", fgColor="FFF2CC")
DUP_RED = PatternFill(fill_type="solid", fgColor="F4CCCC")
DUP_ORANGE = PatternFill(fill_type="solid", fgColor="FCE5CD")
QA_ORANGE = PatternFill(fill_type="solid", fgColor="FCE5CD")
OK_GREEN = PatternFill(fill_type="solid", fgColor="D9EAD3")
HEADER_BLUE = PatternFill(fill_type="solid", fgColor="D9EAF7")

DOMAIN_KEYWORDS = {
    "marketing": [
        "campaign", "brand", "customer", "audience", "market", "solution", "experience",
        "discover", "event", "register", "join us", "learn more", "webinar", "summit",
        "innovation", "enterprise", "cta", "demo", "story", "buyer", "sales",
    ],
    "legal": [
        "agreement", "terms", "privacy", "policy", "consent", "law", "compliance",
        "contract", "liability", "regulation", "notice", "binding", "processing",
    ],
    "technical": [
        "server", "database", "api", "deployment", "instance", "integration", "cloud",
        "compute", "storage", "platform", "kubernetes", "tenant", "authentication",
        "infrastructure", "cluster", "configuration", "endpoint",
    ],
    "ui/software": [
        "click", "save", "cancel", "submit", "sign in", "settings", "next", "back",
        "menu", "dashboard", "select", "search", "apply", "button", "field", "dialog",
    ],
    "support": [
        "ticket", "support", "help", "issue", "troubleshoot", "contact us", "knowledge base",
        "service request", "incident", "resolution", "case",
    ],
    "hr": [
        "employee", "manager", "benefits", "payroll", "hiring", "candidate", "performance review",
        "workforce", "talent", "recruiting", "onboarding",
    ],
    "finance": [
        "invoice", "revenue", "expense", "payment", "budget", "tax", "financial", "ledger",
        "purchase order", "billing", "forecast",
    ],
}

FILE_DOMAIN_HINTS = {
    "marketing": ["mkt", "campaign", "event", "social", "web", "landing"],
    "legal": ["legal", "privacy", "terms", "policy", "dpa"],
    "technical": ["tech", "api", "dev", "cloud", "infra", "docs"],
    "ui/software": ["ui", "screen", "strings", "software", "productui"],
    "support": ["support", "kb", "knowledge", "help", "incident"],
    "hr": ["hr", "employee", "talent", "recruit"],
    "finance": ["finance", "billing", "invoice", "tax"],
}


@dataclass
class SegmentRecord:
    record_id: int
    file_name: str
    tu_index: int
    tuid: str
    source_lang: str
    target_lang: str
    source_text: str
    target_text: str
    creation_date: str
    change_date: str
    creation_id: str
    change_id: str
    domain: str = "general"
    duplicate_type: str = ""
    duplicate_group: str = ""
    glossary_status: str = ""
    glossary_details: str = ""
    qa_status: str = ""
    qa_issues: str = ""
    source_length: int = 0
    target_length: int = 0


class TMXParser:
    @staticmethod
    def _clean_text(text: str) -> str:
        text = text or ""
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    @staticmethod
    def _strip_namespace(tag: str) -> str:
        return tag.split("}", 1)[1] if "}" in tag else tag

    @staticmethod
    def _xml_lang(elem: ET.Element) -> Optional[str]:
        return elem.attrib.get("{http://www.w3.org/XML/1998/namespace}lang") or elem.attrib.get("lang")

    @staticmethod
    def _lang_matches(actual: str, preferred: str) -> bool:
        if not preferred:
            return True
        a = (actual or "").lower().strip()
        p = (preferred or "").lower().strip()
        return a == p or a.startswith(p + "-") or a.startswith(p + "_")

    @staticmethod
    def parse_tmx(
        path: str,
        start_record_id: int = 1,
        preferred_source: str = "",
        preferred_target: str = "",
    ) -> List[SegmentRecord]:
        segments: List[SegmentRecord] = []
        file_name = os.path.basename(path)

        try:
            context = ET.iterparse(path, events=("end",))
        except ET.ParseError as e:
            raise ValueError(f"Could not parse TMX file: {path}\n{e}") from e

        tu_counter = 0
        next_record_id = start_record_id

        for _event, elem in context:
            if TMXParser._strip_namespace(elem.tag) != "tu":
                continue

            tu_counter += 1
            tuid = elem.attrib.get("tuid", "")
            creation_date = elem.attrib.get("creationdate", "")
            change_date = elem.attrib.get("changedate", "")
            creation_id = elem.attrib.get("creationid", "")
            change_id = elem.attrib.get("changeid", "")

            tuvs: List[Tuple[str, str]] = []
            for child in list(elem):
                if TMXParser._strip_namespace(child.tag) != "tuv":
                    continue
                lang = (TMXParser._xml_lang(child) or "").strip().lower()
                seg_text = ""
                for seg in list(child):
                    if TMXParser._strip_namespace(seg.tag) == "seg":
                        seg_text = "".join(seg.itertext())
                        break
                tuvs.append((lang, TMXParser._clean_text(seg_text)))

            source_lang = ""
            target_lang = ""
            source_text = ""
            target_text = ""

            if len(tuvs) >= 2:
                if preferred_source and preferred_target:
                    src = next((x for x in tuvs if TMXParser._lang_matches(x[0], preferred_source)), None)
                    tgt = next((x for x in tuvs if TMXParser._lang_matches(x[0], preferred_target)), None)
                    if src and tgt and src != tgt:
                        source_lang, source_text = src
                        target_lang, target_text = tgt
                    else:
                        source_lang, source_text = tuvs[0]
                        target_lang, target_text = tuvs[1]
                else:
                    source_lang, source_text = tuvs[0]
                    target_lang, target_text = tuvs[1]
            elif len(tuvs) == 1:
                source_lang, source_text = tuvs[0]
                target_lang, target_text = "", ""
            else:
                elem.clear()
                continue

            segments.append(
                SegmentRecord(
                    record_id=next_record_id,
                    file_name=file_name,
                    tu_index=tu_counter,
                    tuid=tuid,
                    source_lang=source_lang,
                    target_lang=target_lang,
                    source_text=source_text,
                    target_text=target_text,
                    creation_date=creation_date,
                    change_date=change_date,
                    creation_id=creation_id,
                    change_id=change_id,
                    source_length=len(source_text),
                    target_length=len(target_text),
                )
            )
            next_record_id += 1
            elem.clear()

        return segments


class TMXRepair:
    LANGUAGE_CODE_MAP = {
        "de": "de-DE", "de-de": "de-DE", "de_de": "de-DE", "ger": "de-DE", "deu": "de-DE",
        "en": "en-US", "en-us": "en-US", "en_us": "en-US", "eng": "en-US",
        "en-uk": "en-GB", "en_uk": "en-GB", "en-gb": "en-GB", "en_gb": "en-GB",
        "fr": "fr-FR", "fr-fr": "fr-FR", "fr_fr": "fr-FR", "fre": "fr-FR", "fra": "fr-FR",
        "fr-ca": "fr-CA", "fr_ca": "fr-CA",
        "es": "es-ES", "es-es": "es-ES", "es_es": "es-ES", "spa": "es-ES",
        "es-mx": "es-MX", "es_mx": "es-MX", "es-ww": "es-WW", "es_ww": "es-WW",
        "it": "it-IT", "it-it": "it-IT", "it_it": "it-IT", "ita": "it-IT",
        "pt": "pt-PT", "pt-pt": "pt-PT", "pt_pt": "pt-PT", "por": "pt-PT",
        "pt-br": "pt-BR", "pt_br": "pt-BR",
        "nl": "nl-NL", "nl-nl": "nl-NL", "nl_nl": "nl-NL", "dut": "nl-NL", "nld": "nl-NL",
        "ja": "ja-JP", "ja-jp": "ja-JP", "ja_jp": "ja-JP", "jpn": "ja-JP",
        "ko": "ko-KR", "ko-kr": "ko-KR", "ko_kr": "ko-KR", "kor": "ko-KR",
        "zh": "zh-CN", "zh-cn": "zh-CN", "zh_cn": "zh-CN", "chi": "zh-CN", "zho": "zh-CN",
        "zh-tw": "zh-TW", "zh_tw": "zh-TW",
        "ar": "ar-SA", "ar-sa": "ar-SA", "ar_sa": "ar-SA", "ara": "ar-SA",
        "pl": "pl-PL", "pl-pl": "pl-PL", "pl_pl": "pl-PL", "pol": "pl-PL",
        "cs": "cs-CZ", "cs-cz": "cs-CZ", "cs_cz": "cs-CZ", "cze": "cs-CZ", "ces": "cs-CZ",
        "sv": "sv-SE", "sv-se": "sv-SE", "sv_se": "sv-SE", "swe": "sv-SE",
        "da": "da-DK", "da-dk": "da-DK", "da_dk": "da-DK", "dan": "da-DK",
        "nb": "nb-NO", "nb-no": "nb-NO", "nb_no": "nb-NO", "no": "nb-NO", "nor": "nb-NO",
        "fi": "fi-FI", "fi-fi": "fi-FI", "fi_fi": "fi-FI", "fin": "fi-FI",
        "tr": "tr-TR", "tr-tr": "tr-TR", "tr_tr": "tr-TR", "tur": "tr-TR",
        "ru": "ru-RU", "ru-ru": "ru-RU", "ru_ru": "ru-RU", "rus": "ru-RU",
    }

    @staticmethod
    def normalize_language_code(lang: str) -> str:
        raw = (lang or "").strip()
        if not raw:
            return ""
        key = raw.lower().replace("_", "-")
        if key in TMXRepair.LANGUAGE_CODE_MAP:
            return TMXRepair.LANGUAGE_CODE_MAP[key]
        parts = key.split("-")
        if len(parts) == 2 and len(parts[0]) == 2 and len(parts[1]) == 2:
            return f"{parts[0].lower()}-{parts[1].upper()}"
        return key

    @staticmethod
    def repair_text(
        text: str,
        trim_spaces: bool = True,
        remove_hidden_chars: bool = True,
        collapse_spaces: bool = True,
        unicode_normalize: bool = False,
    ) -> str:
        text = text or ""
        if unicode_normalize:
            text = unicodedata.normalize("NFC", text)
        if remove_hidden_chars:
            text = text.replace("\xa0", " ")
            text = text.replace("\u200b", "")
            text = text.replace("\ufeff", "")
        if collapse_spaces:
            text = re.sub(r"[ \t]+", " ", text)
            text = re.sub(r"\s+", " ", text)
        if trim_spaces:
            text = text.strip()
        return text

    @staticmethod
    def repair_record(
        r: SegmentRecord,
        normalize_lang_codes: bool = True,
        trim_spaces: bool = True,
        remove_hidden_chars: bool = True,
        collapse_spaces: bool = True,
        unicode_normalize: bool = False,
    ) -> None:
        r.source_text = TMXRepair.repair_text(r.source_text, trim_spaces, remove_hidden_chars, collapse_spaces, unicode_normalize)
        r.target_text = TMXRepair.repair_text(r.target_text, trim_spaces, remove_hidden_chars, collapse_spaces, unicode_normalize)
        if normalize_lang_codes:
            r.source_lang = TMXRepair.normalize_language_code(r.source_lang)
            r.target_lang = TMXRepair.normalize_language_code(r.target_lang)
        else:
            r.source_lang = (r.source_lang or "").strip().replace("_", "-")
            r.target_lang = (r.target_lang or "").strip().replace("_", "-")
        if r.tuid:
            r.tuid = r.tuid.strip()
        r.source_length = len(r.source_text)
        r.target_length = len(r.target_text)

    @staticmethod
    def repair_all(
        records: List[SegmentRecord],
        normalize_lang_codes: bool = True,
        trim_spaces: bool = True,
        remove_hidden_chars: bool = True,
        collapse_spaces: bool = True,
        unicode_normalize: bool = False,
    ) -> int:
        changed = 0
        for r in records:
            before = (r.source_text, r.target_text, r.source_lang, r.target_lang, r.tuid)
            TMXRepair.repair_record(
                r,
                normalize_lang_codes=normalize_lang_codes,
                trim_spaces=trim_spaces,
                remove_hidden_chars=remove_hidden_chars,
                collapse_spaces=collapse_spaces,
                unicode_normalize=unicode_normalize,
            )
            after = (r.source_text, r.target_text, r.source_lang, r.target_lang, r.tuid)
            if before != after:
                changed += 1
        return changed


class DomainClassifier:
    @staticmethod
    def classify(source: str, target: str, file_name: str = "") -> str:
        blob = f"{source} {target}".lower()
        file_blob = (file_name or "").lower()
        scores: Dict[str, int] = defaultdict(int)

        for domain, keywords in DOMAIN_KEYWORDS.items():
            scores[domain] += sum(3 for kw in keywords if kw in blob)
        for domain, hints in FILE_DOMAIN_HINTS.items():
            scores[domain] += sum(2 for hint in hints if hint in file_blob)

        if re.search(r"\b(click|save|cancel|submit|next|back)\b", blob):
            scores["ui/software"] += 3
        if re.search(r"\b(agreement|privacy|policy|liability|consent)\b", blob):
            scores["legal"] += 3
        if re.search(r"\b(cloud|api|database|tenant|kubernetes|endpoint)\b", blob):
            scores["technical"] += 3

        if not scores:
            return "general"
        best_domain, best_score = max(scores.items(), key=lambda x: x[1])
        return best_domain if best_score > 0 else "general"


class DuplicateAnalyzer:
    @staticmethod
    def norm(text: str) -> str:
        text = re.sub(r"\s+", " ", (text or "").strip().lower())
        text = re.sub(r"[\W_]+", "", text, flags=re.UNICODE)
        return text

    @staticmethod
    def _key_exact_pair(r: SegmentRecord) -> Tuple[str, str, str, str]:
        return (r.source_lang, r.target_lang, r.source_text, r.target_text)

    @staticmethod
    def _key_source_lang(r: SegmentRecord) -> Tuple[str, str]:
        return (r.source_lang, r.source_text)

    @staticmethod
    def _key_norm_source_lang(r: SegmentRecord) -> Tuple[str, str]:
        return (r.source_lang, DuplicateAnalyzer.norm(r.source_text))

    @staticmethod
    def apply(records: List[SegmentRecord]) -> None:
        for r in records:
            r.duplicate_type = ""
            r.duplicate_group = ""

        exact_pair_counter = Counter(DuplicateAnalyzer._key_exact_pair(r) for r in records)
        source_to_targets: Dict[Tuple[str, str], set] = defaultdict(set)
        norm_source_counter = Counter(DuplicateAnalyzer._key_norm_source_lang(r) for r in records)
        exact_group_ids: Dict[Tuple[str, str, str, str], str] = {}
        source_group_ids: Dict[Tuple[str, str], str] = {}
        norm_group_ids: Dict[Tuple[str, str], str] = {}

        for r in records:
            source_to_targets[DuplicateAnalyzer._key_source_lang(r)].add(r.target_text)

        exact_idx = 1
        source_idx = 1
        norm_idx = 1

        for r in records:
            exact_key = DuplicateAnalyzer._key_exact_pair(r)
            src_key = DuplicateAnalyzer._key_source_lang(r)
            norm_key = DuplicateAnalyzer._key_norm_source_lang(r)

            if exact_pair_counter[exact_key] > 1:
                if exact_key not in exact_group_ids:
                    exact_group_ids[exact_key] = f"EXACT-{exact_idx:04d}"
                    exact_idx += 1
                r.duplicate_type = "Exact source+target duplicate"
                r.duplicate_group = exact_group_ids[exact_key]

            if len(source_to_targets[src_key]) > 1:
                if src_key not in source_group_ids:
                    source_group_ids[src_key] = f"SOURCE-{source_idx:04d}"
                    source_idx += 1
                r.duplicate_type = "Same source, different target"
                r.duplicate_group = source_group_ids[src_key]
            elif norm_source_counter[norm_key] > 1 and not r.duplicate_type:
                if norm_key not in norm_group_ids:
                    norm_group_ids[norm_key] = f"NORM-{norm_idx:04d}"
                    norm_idx += 1
                r.duplicate_type = "Normalized source duplicate"
                r.duplicate_group = norm_group_ids[norm_key]


class GlossaryEngine:
    def __init__(self) -> None:
        self.terms: List[Dict[str, str]] = []
        self.forbidden_terms: List[Dict[str, str]] = []

    def load_xlsx(self, path: str) -> Tuple[int, str]:
        df = pd.read_excel(path, header=None)
        if df.shape[1] < 2:
            raise ValueError("Glossary XLSX must contain at least two columns: source term and approved target term.")

        self.terms = []
        self.forbidden_terms = []

        for _, row in df.iterrows():
            src = "" if pd.isna(row.iloc[0]) else str(row.iloc[0]).strip()
            tgt = "" if pd.isna(row.iloc[1]) else str(row.iloc[1]).strip()
            domain = "" if df.shape[1] < 3 or pd.isna(row.iloc[2]) else str(row.iloc[2]).strip()
            note = "" if df.shape[1] < 4 or pd.isna(row.iloc[3]) else str(row.iloc[3]).strip()
            status = "" if df.shape[1] < 5 or pd.isna(row.iloc[4]) else str(row.iloc[4]).strip().lower()
            if src and tgt:
                item = {"source": src, "target": tgt, "domain": domain, "note": note}
                self.terms.append(item)
                if status in {"forbidden", "avoid", "deprecated"}:
                    self.forbidden_terms.append(item)

        return len(self.terms), os.path.basename(path)

    @staticmethod
    def _contains_term(text: str, term: str, whole_word: bool = True) -> bool:
        if not text or not term:
            return False
        if whole_word:
            pattern = r"(?<!\w)" + re.escape(term) + r"(?!\w)"
            return re.search(pattern, text, flags=re.IGNORECASE) is not None
        return term.lower() in text.lower()

    @staticmethod
    def _inflection_match(text: str, term: str) -> bool:
        if not text or not term:
            return False
        base = re.escape(term.rstrip())
        pattern = rf"(?<!\w){base}(e|en|er|es|n|s)?(?!\w)"
        return re.search(pattern, text, flags=re.IGNORECASE) is not None

    def apply(self, records: List[SegmentRecord], whole_word: bool = True) -> None:
        for r in records:
            r.glossary_status = ""
            r.glossary_details = ""
        if not self.terms:
            return

        for r in records:
            hits = []
            violations = []
            domain_hints = []

            for item in self.terms:
                src_term = item["source"]
                tgt_term = item["target"]
                domain = item.get("domain", "")
                if self._contains_term(r.source_text, src_term, whole_word=whole_word):
                    hits.append(f"{src_term} -> {tgt_term}")
                    if domain:
                        domain_hints.append(domain)
                    if not (
                        self._contains_term(r.target_text, tgt_term, whole_word=whole_word)
                        or self._inflection_match(r.target_text, tgt_term)
                    ):
                        violations.append(f"Expected '{tgt_term}' for '{src_term}'")

            for item in self.forbidden_terms:
                tgt_term = item["target"]
                if self._contains_term(r.target_text, tgt_term, whole_word=False):
                    violations.append(f"Forbidden/avoid term used: '{tgt_term}'")

            if hits and violations:
                r.glossary_status = "Violation"
                r.glossary_details = "; ".join(violations[:10])
            elif hits:
                r.glossary_status = "Matched"
                r.glossary_details = "; ".join(hits[:10])
            else:
                r.glossary_status = "No term hit"
                r.glossary_details = ""

            if r.domain == "general" and domain_hints:
                r.domain = domain_hints[0]


class BrandProtectionEngine:
    def __init__(self) -> None:
        self.terms: List[Dict[str, str]] = []

    def load_xlsx(self, path: str) -> Tuple[int, str]:
        df = pd.read_excel(path, header=None)
        if df.shape[1] < 2:
            raise ValueError("Do-not-translate XLSX must contain at least two columns: source term and required target representation.")
        self.terms = []
        for _, row in df.iterrows():
            src = "" if pd.isna(row.iloc[0]) else str(row.iloc[0]).strip()
            tgt = "" if pd.isna(row.iloc[1]) else str(row.iloc[1]).strip()
            note = "" if df.shape[1] < 3 or pd.isna(row.iloc[2]) else str(row.iloc[2]).strip()
            if src and tgt:
                self.terms.append({"source": src, "target": tgt, "note": note})
        return len(self.terms), os.path.basename(path)

    @staticmethod
    def _contains(text: str, term: str) -> bool:
        if not text or not term:
            return False
        pattern = r"(?<!\w)" + re.escape(term) + r"(?!\w)"
        return re.search(pattern, text, flags=re.IGNORECASE) is not None

    def issues_for_record(self, r: SegmentRecord) -> List[str]:
        issues: List[str] = []
        for item in self.terms:
            src = item["source"]
            required = item["target"]
            if self._contains(r.source_text, src) and not self._contains(r.target_text, required):
                issues.append(f"Brand protection: expected '{required}' for '{src}'")
        return issues


class QAEngine:
    PLACEHOLDER_PATTERNS = [
        r"\{\d+\}",
        r"\{[A-Za-z0-9_]+\}",
        r"%s",
        r"%d",
        r"\$\{[A-Za-z0-9_]+\}",
        r"<[^>]+>",
    ]

    @staticmethod
    def _extract_placeholders(text: str) -> List[str]:
        found = []
        for patt in QAEngine.PLACEHOLDER_PATTERNS:
            found.extend(re.findall(patt, text or ""))
        return sorted(found)

    @staticmethod
    def _numbers(text: str) -> List[str]:
        return re.findall(r"\d+(?:[\.,]\d+)?", text or "")

    @staticmethod
    def _end_punct(text: str) -> str:
        text = (text or "").strip()
        return text[-1] if text and text[-1] in ".,:;!?" else ""

    @staticmethod
    def _bracket_counts(text: str) -> Dict[str, int]:
        return {
            "(": (text or "").count("("),
            ")": (text or "").count(")"),
            "[": (text or "").count("["),
            "]": (text or "").count("]"),
            "{": (text or "").count("{"),
            "}": (text or "").count("}"),
        }

    @staticmethod
    def _quote_counts(text: str) -> Dict[str, int]:
        return {q: (text or "").count(q) for q in ['"', "'", "“", "”", "‘", "’"]}

    @staticmethod
    def _starts_upper(text: str) -> Optional[bool]:
        text = (text or "").strip()
        for ch in text:
            if ch.isalpha():
                return ch.isupper()
        return None

    @staticmethod
    def _has_unbalanced_html_tags(text: str) -> bool:
        text = text or ""
        tag_pattern = re.compile(r"<\s*(/?)\s*([A-Za-z][A-Za-z0-9:_-]*)(?:\s[^<>]*)?>")
        stack: List[str] = []
        self_closing = {"br", "hr", "img", "input", "meta", "link", "area", "base", "col", "embed", "param", "source", "track", "wbr"}
        for match in tag_pattern.finditer(text):
            closing, name = match.group(1), match.group(2).lower()
            full = match.group(0)
            if full.endswith("/>") or name in self_closing:
                continue
            if closing:
                if not stack or stack[-1] != name:
                    return True
                stack.pop()
            else:
                stack.append(name)
        return bool(stack)

    @staticmethod
    def _normalized_equal(a: str, b: str) -> bool:
        return re.sub(r"\s+", " ", (a or "").strip()) == re.sub(r"\s+", " ", (b or "").strip())

    @staticmethod
    def _german_micro_issues(target: str) -> List[str]:
        issues: List[str] = []
        t = target or ""
        if re.search(r"\s+[.,:;!?]", t):
            issues.append("German micro QA: space before punctuation")
        if re.search(r"\b(\w+)\s+\1\b", t, flags=re.IGNORECASE):
            issues.append("German micro QA: repeated word")
        if re.search(r'(^|\s)["][A-Za-zÄÖÜäöüß]', t) or re.search(r'[A-Za-zÄÖÜäöüß]["](\s|$|[.,:;!?])', t):
            issues.append("German micro QA: English straight quotes used")
        if re.search(r"\d\s+%", t):
            issues.append("German micro QA: space before percent sign")
        if re.search(r"\d+\.\d+", t):
            issues.append("German micro QA: possible decimal point instead of decimal comma")
        return issues

    @staticmethod
    def apply(records: List[SegmentRecord], brand_engine: Optional["BrandProtectionEngine"] = None) -> None:
        for r in records:
            issues = []
            s = r.source_text or ""
            t = r.target_text or ""

            if not t.strip():
                issues.append("Missing target")
            if s.strip() and t.strip() and s.strip() == t.strip():
                issues.append("Source equals target")
            if s.strip() and t.strip() and QAEngine._normalized_equal(s, t):
                issues.append("Target equals source / possible untranslated segment")
            if "  " in t:
                issues.append("Double spaces in target")
            if "\xa0" in t:
                issues.append("Non-breaking space in target")
            if s[:1].isspace() != t[:1].isspace() or s[-1:].isspace() != t[-1:].isspace():
                issues.append("Leading/trailing space mismatch")
            if QAEngine._end_punct(s) != QAEngine._end_punct(t):
                if QAEngine._end_punct(s) or QAEngine._end_punct(t):
                    issues.append("Ending punctuation mismatch")
            if QAEngine._numbers(s) != QAEngine._numbers(t):
                issues.append("Number mismatch")
            if QAEngine._extract_placeholders(s) != QAEngine._extract_placeholders(t):
                issues.append("Placeholder/tag mismatch")
            if QAEngine._bracket_counts(s) != QAEngine._bracket_counts(t):
                issues.append("Bracket mismatch")
            if QAEngine._quote_counts(s) != QAEngine._quote_counts(t):
                issues.append("Quote mismatch")
            if len(s) > 0:
                ratio = len(t) / max(1, len(s))
                if ratio < 0.35 or ratio > 2.8:
                    issues.append("Suspicious length ratio")
            if re.search(r"\b(\w+)\s+\1\b", t, flags=re.IGNORECASE):
                issues.append("Repeated word in target")
            src_upper = QAEngine._starts_upper(s)
            tgt_upper = QAEngine._starts_upper(t)
            if src_upper is not None and tgt_upper is not None and src_upper != tgt_upper:
                issues.append("Capitalization mismatch")
            if re.search(r"\s+[.,:;!?]", t):
                issues.append("Space before punctuation")
            if QAEngine._has_unbalanced_html_tags(s):
                issues.append("Unbalanced HTML/XML tag in source")
            if QAEngine._has_unbalanced_html_tags(t):
                issues.append("Unbalanced HTML/XML tag in target")
            if (r.target_lang or "").lower().startswith("de"):
                issues.extend(QAEngine._german_micro_issues(t))
            if brand_engine is not None:
                issues.extend(brand_engine.issues_for_record(r))
            if r.glossary_status == "Violation":
                issues.append("Glossary violation")

            r.qa_status = "Issues" if issues else "OK"
            r.qa_issues = "; ".join(issues)


class StatsEngine:
    @staticmethod
    def build(records: List[SegmentRecord]) -> Dict[str, object]:
        stats: Dict[str, object] = {}
        stats["total_segments"] = len(records)
        stats["unique_sources"] = len({r.source_text for r in records})
        stats["unique_targets"] = len({r.target_text for r in records})
        stats["duplicate_segments"] = sum(1 for r in records if r.duplicate_type)
        stats["inconsistent_sources"] = sum(1 for r in records if r.duplicate_type == "Same source, different target")
        stats["qa_issue_count"] = sum(1 for r in records if r.qa_status == "Issues")
        stats["glossary_violations"] = sum(1 for r in records if r.glossary_status == "Violation")
        stats["average_source_length"] = round(sum(r.source_length for r in records) / max(1, len(records)), 2)
        stats["average_target_length"] = round(sum(r.target_length for r in records) / max(1, len(records)), 2)
        stats["language_pairs"] = Counter(f"{r.source_lang} > {r.target_lang}" for r in records)
        stats["domain_distribution"] = Counter(r.domain for r in records)
        stats["per_file_counts"] = Counter(r.file_name for r in records)
        stats["top_repeated_sources"] = Counter(r.source_text for r in records).most_common(20)
        return stats


class MergeResolver:
    @staticmethod
    def _parse_change_date(value: str) -> datetime:
        value = (value or "").strip()
        if not value:
            return datetime.min
        for fmt in ["%Y%m%dT%H%M%SZ", "%Y%m%dT%H%M%S", "%Y%m%d", "%Y-%m-%d"]:
            try:
                return datetime.strptime(value, fmt)
            except Exception:
                continue
        return datetime.min

    @staticmethod
    def choose_best(records: List[SegmentRecord], strategy: str) -> SegmentRecord:
        if not records:
            raise ValueError("No records supplied to choose_best")
        if strategy == "keep newest translation":
            return max(
                records,
                key=lambda r: (
                    MergeResolver._parse_change_date(r.change_date),
                    MergeResolver._parse_change_date(r.creation_date),
                    len(r.target_text or ""),
                ),
            )
        if strategy == "keep longest translation":
            return max(records, key=lambda r: (len(r.target_text or ""), len(r.source_text or "")))
        if strategy == "keep glossary-compliant translation":
            sorted_records = sorted(
                records,
                key=lambda r: (
                    1 if r.glossary_status == "Matched" else 0,
                    0 if r.glossary_status == "Violation" else 1,
                    len(r.target_text or ""),
                    MergeResolver._parse_change_date(r.change_date),
                ),
                reverse=True,
            )
            return sorted_records[0]
        return records[0]

    @staticmethod
    def merge_records(records: List[SegmentRecord], strategy: str) -> Tuple[List[SegmentRecord], int]:
        grouped: Dict[Tuple[str, str, str], List[SegmentRecord]] = defaultdict(list)
        for r in records:
            grouped[(r.source_lang, r.target_lang, r.source_text)].append(r)
        merged: List[SegmentRecord] = []
        removed = 0
        for group in grouped.values():
            if len(group) == 1:
                merged.append(group[0])
            else:
                best = MergeResolver.choose_best(group, strategy)
                merged.append(best)
                removed += len(group) - 1
        merged.sort(key=lambda r: (r.file_name, r.tu_index, r.record_id))
        return merged, removed


class TMXExporter:
    @staticmethod
    def export(records: List[SegmentRecord], output_path: str, maker: str = MAKER_LINE) -> None:
        root = ET.Element("tmx", version="1.4")
        ET.SubElement(
            root,
            "header",
            {
                "creationtool": maker,
                "creationtoolversion": "2.2",
                "segtype": "sentence",
                "adminlang": "en-us",
                "srclang": records[0].source_lang if records else "en",
                "datatype": "PlainText",
            },
        )
        body = ET.SubElement(root, "body")
        for idx, r in enumerate(records, start=1):
            tu_attrib = {"tuid": r.tuid or str(idx)}
            if r.creation_date:
                tu_attrib["creationdate"] = r.creation_date
            if r.change_date:
                tu_attrib["changedate"] = r.change_date
            if r.creation_id:
                tu_attrib["creationid"] = r.creation_id
            if r.change_id:
                tu_attrib["changeid"] = r.change_id
            tu = ET.SubElement(body, "tu", tu_attrib)
            tuv_src = ET.SubElement(
                tu,
                "tuv",
                {"{http://www.w3.org/XML/1998/namespace}lang": r.source_lang or "en"},
            )
            ET.SubElement(tuv_src, "seg").text = r.source_text
            tuv_tgt = ET.SubElement(
                tu,
                "tuv",
                {"{http://www.w3.org/XML/1998/namespace}lang": r.target_lang or "de"},
            )
            ET.SubElement(tuv_tgt, "seg").text = r.target_text
        ET.ElementTree(root).write(output_path, encoding="utf-8", xml_declaration=True)


class ChartBuilder:
    @staticmethod
    def create_charts(records: List[SegmentRecord], output_dir: str) -> List[str]:
        os.makedirs(output_dir, exist_ok=True)
        paths: List[str] = []

        domain_counts = Counter(r.domain for r in records)
        if domain_counts:
            path = os.path.join(output_dir, "domain_distribution.png")
            fig = plt.figure(figsize=(10, 5))
            ax = fig.add_subplot(111)
            ax.bar(list(domain_counts.keys()), list(domain_counts.values()))
            ax.set_title("Domain Distribution")
            ax.set_xlabel("Domain")
            ax.set_ylabel("Count")
            plt.xticks(rotation=35, ha="right")
            plt.tight_layout(pad=2.0)
            fig.savefig(path, dpi=150, bbox_inches="tight")
            plt.close(fig)
            paths.append(path)

        issue_counts = {
            "Duplicates": sum(1 for r in records if r.duplicate_type),
            "Glossary Violations": sum(1 for r in records if r.glossary_status == "Violation"),
            "QA Issues": sum(1 for r in records if r.qa_status == "Issues"),
        }
        path = os.path.join(output_dir, "issue_overview.png")
        fig = plt.figure(figsize=(8, 5))
        ax = fig.add_subplot(111)
        ax.bar(list(issue_counts.keys()), list(issue_counts.values()))
        ax.set_title("Issue Overview")
        ax.set_ylabel("Count")
        plt.tight_layout(pad=2.0)
        fig.savefig(path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        paths.append(path)

        return paths


class XLSXExporter:
    COLUMNS = [
        "Record ID", "File", "TU Index", "TU ID", "Source Lang", "Target Lang", "Source", "Target",
        "Domain", "Duplicate Type", "Duplicate Group", "Glossary Status", "Glossary Details",
        "QA Status", "QA Issues", "Source Length", "Target Length",
        "Creation Date", "Change Date", "Creation ID", "Change ID",
    ]

    @staticmethod
    def _autofit(ws) -> None:
        for col in ws.columns:
            max_len = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[letter].width = min(max_len + 2, 80)

    @staticmethod
    def _style_header(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_BLUE
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    @staticmethod
    def _append_record(ws, r: SegmentRecord) -> None:
        ws.append([
            r.record_id, r.file_name, r.tu_index, r.tuid, r.source_lang, r.target_lang,
            r.source_text, r.target_text, r.domain, r.duplicate_type, r.duplicate_group,
            r.glossary_status, r.glossary_details, r.qa_status, r.qa_issues,
            r.source_length, r.target_length, r.creation_date, r.change_date,
            r.creation_id, r.change_id,
        ])

    @staticmethod
    def export(records: List[SegmentRecord], stats: Dict[str, object], output_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Segments"
        ws.append(XLSXExporter.COLUMNS)
        for r in records:
            XLSXExporter._append_record(ws, r)
        XLSXExporter._style_header(ws)
        ws.freeze_panes = "A2"
        XLSXExporter._autofit(ws)

        for row in range(2, ws.max_row + 1):
            dup_type = ws.cell(row=row, column=10).value or ""
            glossary_status = ws.cell(row=row, column=12).value or ""
            qa_status = ws.cell(row=row, column=14).value or ""
            fill = None
            if dup_type == "Exact source+target duplicate":
                fill = DUP_YELLOW
            elif dup_type == "Same source, different target":
                fill = DUP_RED
            elif dup_type == "Normalized source duplicate":
                fill = DUP_ORANGE
            elif qa_status == "Issues" or glossary_status == "Violation":
                fill = QA_ORANGE
            if fill:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=c).fill = fill
            else:
                ws.cell(row=row, column=14).fill = OK_GREEN

        for title, filter_fn in [
            ("Duplicates", lambda r: bool(r.duplicate_type)),
            ("Glossary", lambda r: r.glossary_status in {"Matched", "Violation"}),
            ("QA", lambda r: r.qa_status == "Issues"),
        ]:
            t_ws = wb.create_sheet(title)
            t_ws.append(XLSXExporter.COLUMNS)
            for r in records:
                if filter_fn(r):
                    XLSXExporter._append_record(t_ws, r)
            XLSXExporter._style_header(t_ws)
            t_ws.freeze_panes = "A2"
            XLSXExporter._autofit(t_ws)

        stats_ws = wb.create_sheet("Statistics")
        stats_ws.append(["Metric", "Value"])
        XLSXExporter._style_header(stats_ws)
        basic = [
            ("Tool Maker", MAKER_LINE),
            ("Total Segments", stats.get("total_segments", 0)),
            ("Unique Sources", stats.get("unique_sources", 0)),
            ("Unique Targets", stats.get("unique_targets", 0)),
            ("Duplicate Segments", stats.get("duplicate_segments", 0)),
            ("Inconsistent Sources", stats.get("inconsistent_sources", 0)),
            ("QA Issue Count", stats.get("qa_issue_count", 0)),
            ("Glossary Violations", stats.get("glossary_violations", 0)),
            ("Average Source Length", stats.get("average_source_length", 0)),
            ("Average Target Length", stats.get("average_target_length", 0)),
        ]
        for row in basic:
            stats_ws.append(list(row))
        stats_ws.append([])
        stats_ws.append(["Language Pairs", "Count"])
        for k, v in stats.get("language_pairs", {}).items():
            stats_ws.append([k, v])
        stats_ws.append([])
        stats_ws.append(["Domain Distribution", "Count"])
        for k, v in stats.get("domain_distribution", {}).items():
            stats_ws.append([k, v])
        stats_ws.append([])
        stats_ws.append(["Per File Counts", "Count"])
        for k, v in stats.get("per_file_counts", {}).items():
            stats_ws.append([k, v])
        stats_ws.append([])
        stats_ws.append(["Top Repeated Sources", "Count"])
        for k, v in stats.get("top_repeated_sources", []):
            stats_ws.append([k, v])
        stats_ws.freeze_panes = "A2"
        XLSXExporter._autofit(stats_ws)

        chart_dir = os.path.join(os.path.dirname(output_path) or ".", "tmx_analyzer_charts")
        chart_paths = ChartBuilder.create_charts(records, chart_dir)
        if chart_paths:
            charts_ws = wb.create_sheet("Charts")
            charts_ws.column_dimensions["A"].width = 10
            row_anchor = 2
            for chart_path in chart_paths:
                try:
                    img = XLImage(chart_path)
                    charts_ws.add_image(img, f"B{row_anchor}")
                    row_anchor += 28
                except Exception:
                    charts_ws.append([chart_path])
                    row_anchor += 3

        wb.save(output_path)


# ============================================================
# Streamlit app layer
# ============================================================

DISPLAY_COLUMNS = [
    "record_id", "file_name", "tu_index", "source_lang", "target_lang",
    "source_text", "target_text", "domain", "duplicate_type",
    "duplicate_group", "glossary_status", "glossary_details", "qa_status", "qa_issues",
]

DISPLAY_RENAME = {
    "record_id": "Record ID", "file_name": "File", "tu_index": "TU Index",
    "source_lang": "Source Lang", "target_lang": "Target Lang",
    "source_text": "Source", "target_text": "Target", "domain": "Domain",
    "duplicate_type": "Duplicate Type", "duplicate_group": "Duplicate Group",
    "glossary_status": "Glossary Status", "glossary_details": "Glossary Details",
    "qa_status": "QA Status", "qa_issues": "QA Issues",
}

MERGE_POLICIES = [
    "keep newest translation",
    "keep longest translation",
    "keep glossary-compliant translation",
]


def init_state() -> None:
    defaults = {
        "records": [],
        "stats": {},
        "glossary_engine": GlossaryEngine(),
        "brand_engine": BrandProtectionEngine(),
        "logs": ["Application started."],
        "glossary_name": "",
        "brand_name": "",
        "current_page": 1,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def log(message: str) -> None:
    timestamp = datetime.now().strftime("%H:%M:%S")
    st.session_state.logs.append(f"[{timestamp}] {message}")


def save_upload_to_temp(uploaded_file, suffix: str) -> str:
    real_suffix = os.path.splitext(uploaded_file.name)[1] or suffix
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=real_suffix)
    tmp.write(uploaded_file.getvalue())
    tmp.close()
    return tmp.name


def load_glossary_from_upload(uploaded_file) -> Tuple[int, str]:
    path = save_upload_to_temp(uploaded_file, ".xlsx")
    return st.session_state.glossary_engine.load_xlsx(path)


def load_brand_rules_from_upload(uploaded_file) -> Tuple[int, str]:
    path = save_upload_to_temp(uploaded_file, ".xlsx")
    return st.session_state.brand_engine.load_xlsx(path)


def recalculate_all(records: List[SegmentRecord], whole_word: bool = True) -> Dict[str, object]:
    for r in records:
        if not r.domain or r.domain == "general":
            r.domain = DomainClassifier.classify(r.source_text, r.target_text, r.file_name)
        r.source_length = len(r.source_text or "")
        r.target_length = len(r.target_text or "")
    DuplicateAnalyzer.apply(records)
    st.session_state.glossary_engine.apply(records, whole_word)
    QAEngine.apply(records, st.session_state.get("brand_engine"))
    return StatsEngine.build(records)


def records_to_dataframe(records: List[SegmentRecord]) -> pd.DataFrame:
    return pd.DataFrame([{
        "record_id": r.record_id,
        "file_name": r.file_name,
        "tu_index": r.tu_index,
        "tuid": r.tuid,
        "source_lang": r.source_lang,
        "target_lang": r.target_lang,
        "source_text": r.source_text,
        "target_text": r.target_text,
        "domain": r.domain,
        "duplicate_type": r.duplicate_type,
        "duplicate_group": r.duplicate_group,
        "glossary_status": r.glossary_status,
        "glossary_details": r.glossary_details,
        "qa_status": r.qa_status,
        "qa_issues": r.qa_issues,
        "source_length": r.source_length,
        "target_length": r.target_length,
        "creation_date": r.creation_date,
        "change_date": r.change_date,
        "creation_id": r.creation_id,
        "change_id": r.change_id,
    } for r in records])


def filter_records(
    records: List[SegmentRecord],
    search: str = "",
    domain: str = "All",
    duplicate_type: str = "All",
    glossary_status: str = "All",
    qa_status: str = "All",
    only_duplicates: bool = False,
    only_glossary: bool = False,
    only_qa: bool = False,
) -> List[SegmentRecord]:
    search_l = (search or "").lower().strip()
    result = []
    for r in records:
        if only_duplicates and not r.duplicate_type:
            continue
        if only_glossary and r.glossary_status not in {"Matched", "Violation"}:
            continue
        if only_qa and r.qa_status != "Issues":
            continue
        if domain != "All" and r.domain != domain:
            continue
        if duplicate_type != "All" and r.duplicate_type != duplicate_type:
            continue
        if glossary_status != "All" and r.glossary_status != glossary_status:
            continue
        if qa_status != "All" and r.qa_status != qa_status:
            continue
        if search_l:
            blob = " ".join([
                r.file_name, r.source_lang, r.target_lang, r.source_text, r.target_text,
                r.domain, r.duplicate_type, r.duplicate_group, r.glossary_status,
                r.glossary_details, r.qa_status, r.qa_issues
            ]).lower()
            if search_l not in blob:
                continue
        result.append(r)
    return result


def show_record_table(records: List[SegmentRecord], key: str, editable: bool = False) -> pd.DataFrame:
    df = records_to_dataframe(records)
    if df.empty:
        st.info("No records to display.")
        return df

    view_cols = [c for c in DISPLAY_COLUMNS if c in df.columns]
    display_df = df[view_cols].rename(columns=DISPLAY_RENAME)
    st.caption(f"Showing {len(display_df):,} record(s).")

    if editable:
        return st.data_editor(
            display_df,
            key=key,
            use_container_width=True,
            height=520,
            num_rows="fixed",
            disabled=[
                "Record ID", "File", "TU Index", "Duplicate Type", "Duplicate Group",
                "Glossary Status", "Glossary Details", "QA Status", "QA Issues",
            ],
        )

    st.dataframe(display_df, use_container_width=True, height=520)
    return display_df


def export_xlsx_bytes(records: List[SegmentRecord], stats: Dict[str, object]) -> bytes:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    XLSXExporter.export(records, stats, tmp.name)
    with open(tmp.name, "rb") as f:
        return f.read()


def export_tmx_bytes(records: List[SegmentRecord]) -> bytes:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".tmx")
    tmp.close()
    TMXExporter.export(records, tmp.name)
    with open(tmp.name, "rb") as f:
        return f.read()


def build_overview(stats: Dict[str, object]) -> str:
    if not stats:
        return "No analysis has been run yet."

    brand_count = len(st.session_state.get("brand_engine", BrandProtectionEngine()).terms) if "brand_engine" in st.session_state else 0
    lines = [
        f"Tool Maker: {MAKER_LINE}",
        f"Brand Protection Rules Loaded: {brand_count}",
        f"Total Segments: {stats.get('total_segments', 0)}",
        f"Unique Sources: {stats.get('unique_sources', 0)}",
        f"Unique Targets: {stats.get('unique_targets', 0)}",
        f"Duplicate Segments: {stats.get('duplicate_segments', 0)}",
        f"Inconsistent Sources: {stats.get('inconsistent_sources', 0)}",
        f"QA Issue Count: {stats.get('qa_issue_count', 0)}",
        f"Glossary Violations: {stats.get('glossary_violations', 0)}",
        f"Average Source Length: {stats.get('average_source_length', 0)}",
        f"Average Target Length: {stats.get('average_target_length', 0)}",
        "",
        "Language Pairs:",
    ]
    for k, v in stats.get("language_pairs", {}).items():
        lines.append(f"- {k}: {v}")

    lines.append("")
    lines.append("Domain Distribution:")
    for k, v in stats.get("domain_distribution", {}).items():
        lines.append(f"- {k}: {v}")

    lines.append("")
    lines.append("Per File Counts:")
    for k, v in stats.get("per_file_counts", {}).items():
        lines.append(f"- {k}: {v}")

    lines.append("")
    lines.append("Top Repeated Sources:")
    for k, v in stats.get("top_repeated_sources", []):
        lines.append(f"- {k}: {v}")

    return "\n".join(lines)


def render_charts(records: List[SegmentRecord]) -> None:
    if not records:
        return

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Domain Distribution")
        domain_counts = Counter(r.domain for r in records)
        if domain_counts:
            fig = plt.figure(figsize=(8, 4))
            ax = fig.add_subplot(111)
            ax.bar(list(domain_counts.keys()), list(domain_counts.values()))
            ax.set_xlabel("Domain")
            ax.set_ylabel("Count")
            plt.xticks(rotation=35, ha="right")
            plt.tight_layout()
            st.pyplot(fig)
            plt.close(fig)

    with col2:
        st.subheader("Issue Overview")
        issue_counts = {
            "Duplicates": sum(1 for r in records if r.duplicate_type),
            "Glossary Violations": sum(1 for r in records if r.glossary_status == "Violation"),
            "QA Issues": sum(1 for r in records if r.qa_status == "Issues"),
        }
        fig = plt.figure(figsize=(8, 4))
        ax = fig.add_subplot(111)
        ax.bar(list(issue_counts.keys()), list(issue_counts.values()))
        ax.set_ylabel("Count")
        plt.xticks(rotation=20, ha="right")
        plt.tight_layout()
        st.pyplot(fig)
        plt.close(fig)



def apply_selected_repairs(
    records: List[SegmentRecord],
    whole_word: bool,
    normalize_lang_codes: bool,
    trim_spaces: bool,
    remove_hidden_chars: bool,
    collapse_spaces: bool,
    unicode_normalize: bool,
) -> int:
    changed = TMXRepair.repair_all(
        records,
        normalize_lang_codes=normalize_lang_codes,
        trim_spaces=trim_spaces,
        remove_hidden_chars=remove_hidden_chars,
        collapse_spaces=collapse_spaces,
        unicode_normalize=unicode_normalize,
    )
    st.session_state.stats = recalculate_all(records, whole_word)
    return changed

def run_analysis(uploaded_tmx_files, preferred_source: str, preferred_target: str, whole_word: bool) -> None:
    progress = st.progress(0)
    status = st.empty()

    status.info("Parsing TMX files...")
    all_records: List[SegmentRecord] = []
    next_id = 1

    for idx, uploaded_file in enumerate(uploaded_tmx_files, start=1):
        path = save_upload_to_temp(uploaded_file, ".tmx")
        recs = TMXParser.parse_tmx(
            path,
            next_id,
            preferred_source.strip().lower(),
            preferred_target.strip().lower(),
        )

        for r in recs:
            r.file_name = uploaded_file.name

        all_records.extend(recs)
        if recs:
            next_id = max(r.record_id for r in all_records) + 1

        log(f"Parsed {uploaded_file.name}: {len(recs)} segments")
        progress.progress(int((idx / max(1, len(uploaded_tmx_files))) * 35))

    status.info("Classifying, checking duplicates, glossary, brand rules, and QA...")
    st.session_state.records = all_records
    st.session_state.stats = recalculate_all(st.session_state.records, whole_word)

    progress.progress(100)
    status.success("Analysis completed.")
    log(f"Analysis completed. Records: {len(st.session_state.records)}")


def apply_table_edits(edited_df: pd.DataFrame, records: List[SegmentRecord], whole_word: bool) -> None:
    if edited_df.empty:
        return

    reverse = {v: k for k, v in DISPLAY_RENAME.items()}
    internal = edited_df.rename(columns=reverse)
    by_id = {r.record_id: r for r in records}

    for _, row in internal.iterrows():
        rid = int(row["record_id"])
        r = by_id.get(rid)
        if not r:
            continue

        r.source_lang = str(row.get("source_lang", r.source_lang) or "").strip()
        r.target_lang = str(row.get("target_lang", r.target_lang) or "").strip()
        r.source_text = str(row.get("source_text", r.source_text) or "").strip()
        r.target_text = str(row.get("target_text", r.target_text) or "").strip()
        r.domain = str(row.get("domain", r.domain) or "general").strip() or "general"

    st.session_state.stats = recalculate_all(records, whole_word)
    log("Saved table edits and recalculated analysis.")


def app() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="🧠", layout="wide")
    init_state()

    st.title("TMX Analyzer – Phase 2 Advanced")
    st.caption(MAKER_LINE)

    with st.sidebar:
        st.header("Project Setup")

        uploaded_tmx_files = st.file_uploader(
            "Add TMX file(s)",
            type=["tmx"],
            accept_multiple_files=True,
            help="Upload one or more TMX files for analysis or merge.",
        )

        if uploaded_tmx_files:
            with st.expander("Loaded TMX files", expanded=False):
                for i, f in enumerate(uploaded_tmx_files, start=1):
                    st.write(f"{i}. {f.name}")

        glossary_file = st.file_uploader("Load Glossary XLSX", type=["xlsx"])
        if glossary_file is not None and st.button("Load Glossary", use_container_width=True):
            try:
                count, name = load_glossary_from_upload(glossary_file)
                st.session_state.glossary_name = name
                st.success(f"Glossary loaded: {count} terms")
                log(f"Loaded glossary '{name}' with {count} valid term pairs.")
            except Exception as e:
                st.error(f"Glossary error: {e}")
                log(f"Glossary error: {e}")

        brand_file = st.file_uploader(
            "Load Do Not Translate / Brand Protection XLSX",
            type=["xlsx"],
            help="Column A = protected source term. Column B = required target representation.",
        )
        if brand_file is not None and st.button("Load Brand Rules", use_container_width=True):
            try:
                count, name = load_brand_rules_from_upload(brand_file)
                st.session_state.brand_name = name
                st.success(f"Brand protection rules loaded: {count} terms")
                log(f"Loaded brand protection file '{name}' with {count} rules.")
                if st.session_state.records:
                    st.session_state.stats = recalculate_all(st.session_state.records, True)
                    st.rerun()
            except Exception as e:
                st.error(f"Brand protection error: {e}")
                log(f"Brand protection error: {e}")

        st.divider()
        st.header("Analysis Options")

        preferred_source = st.text_input("Preferred source language", placeholder="e.g. en or en-us")
        preferred_target = st.text_input("Preferred target language", placeholder="e.g. de or de-de")
        whole_word = st.checkbox("Glossary whole-word matching", value=True)
        merge_policy = st.selectbox("Merge policy", MERGE_POLICIES)

        st.divider()
        st.header("Auto Repair Options")
        st.caption("These options change loaded records. Riskier items are only flagged in QA.")
        repair_lang_codes = st.checkbox("Normalize language codes", value=True)
        repair_trim = st.checkbox("Trim leading/trailing spaces", value=True)
        repair_hidden = st.checkbox("Remove hidden characters / NBSP", value=True)
        repair_collapse = st.checkbox("Collapse multiple spaces", value=True)
        repair_unicode = st.checkbox("Normalize Unicode to NFC", value=False)

        if st.button("Run Selected Auto Repairs", use_container_width=True):
            if not st.session_state.records:
                st.warning("Please run the analysis first.")
            else:
                changed = apply_selected_repairs(
                    st.session_state.records,
                    whole_word,
                    normalize_lang_codes=repair_lang_codes,
                    trim_spaces=repair_trim,
                    remove_hidden_chars=repair_hidden,
                    collapse_spaces=repair_collapse,
                    unicode_normalize=repair_unicode,
                )
                log(f"Applied selected auto repairs. Updated {changed} record(s).")
                st.success(f"Selected auto repairs completed. Updated {changed} record(s).")
                st.rerun()

        st.divider()
        st.header("Analysis Actions")

        if st.button("Run Analysis / Merge", type="primary", use_container_width=True):
            if not uploaded_tmx_files:
                st.warning("Please upload at least one TMX file first.")
            else:
                try:
                    run_analysis(uploaded_tmx_files, preferred_source, preferred_target, whole_word)
                    st.rerun()
                except Exception as e:
                    st.error(f"Analysis error: {e}")
                    log(f"Analysis error: {e}")

        if st.button("Apply Merge Policy Now", use_container_width=True):
            if not st.session_state.records:
                st.warning("Please run the analysis first.")
            else:
                merged, removed = MergeResolver.merge_records(st.session_state.records, merge_policy)
                st.session_state.records = merged
                st.session_state.stats = recalculate_all(st.session_state.records, whole_word)
                log(f"Applied merge policy '{merge_policy}'. Removed {removed} overlapping record(s).")
                st.success(f"Removed {removed} overlapping record(s).")
                st.rerun()

        if st.button("Run TMX Repair Mode", use_container_width=True):
            if not st.session_state.records:
                st.warning("Please run the analysis first.")
            else:
                changed = apply_selected_repairs(
                    st.session_state.records,
                    whole_word,
                    normalize_lang_codes=repair_lang_codes,
                    trim_spaces=repair_trim,
                    remove_hidden_chars=repair_hidden,
                    collapse_spaces=repair_collapse,
                    unicode_normalize=repair_unicode,
                )
                log(f"Repair mode completed with selected options. Updated {changed} record(s).")
                st.success(f"Repair mode completed. Updated {changed} record(s).")
                st.rerun()

        if st.button("Delete All Exact Duplicates Except First", use_container_width=True):
            records = st.session_state.records
            if not records:
                st.warning("Please run the analysis first.")
            else:
                seen = set()
                kept = []
                removed = 0

                for r in records:
                    key = (r.source_lang, r.target_lang, r.source_text, r.target_text)
                    if r.duplicate_type == "Exact source+target duplicate" and key in seen:
                        removed += 1
                        continue
                    seen.add(key)
                    kept.append(r)

                st.session_state.records = kept
                st.session_state.stats = recalculate_all(st.session_state.records, whole_word)
                log(f"Deleted {removed} exact duplicate record(s), keeping first occurrences.")
                st.success(f"Deleted {removed} exact duplicate record(s).")
                st.rerun()

        st.divider()
        st.header("Export")

        if st.session_state.records:
            st.download_button(
                "Export Analysis to XLSX",
                data=export_xlsx_bytes(st.session_state.records, st.session_state.stats),
                file_name="tmx_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            st.download_button(
                "Save Current Records as TMX",
                data=export_tmx_bytes(st.session_state.records),
                file_name="cleaned_merged.tmx",
                mime="application/xml",
                use_container_width=True,
            )
        else:
            st.caption("Run an analysis before exporting.")

    stats = st.session_state.stats
    records = st.session_state.records

    if st.session_state.get("brand_name"):
        st.caption(f"Loaded brand protection file: {st.session_state.brand_name}")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Segments", stats.get("total_segments", 0))
    c2.metric("Duplicate Segments", stats.get("duplicate_segments", 0))
    c3.metric("QA Issues", stats.get("qa_issue_count", 0))
    c4.metric("Glossary Violations", stats.get("glossary_violations", 0))

    tab_overview, tab_segments, tab_duplicates, tab_glossary, tab_qa, tab_resolver, tab_logs = st.tabs([
        "Overview", "Segments", "Duplicates", "Glossary", "QA", "Duplicate Resolver", "Logs"
    ])

    with tab_overview:
        st.subheader("Overview")
        st.text_area("Statistics", build_overview(stats), height=340)
        render_charts(records)

    with tab_segments:
        st.subheader("Segments")
        if records:
            col_a, col_b, col_c = st.columns([2, 1, 1])
            with col_a:
                search = st.text_input("Search segments", key="segment_search")
            domains = sorted({r.domain for r in records if r.domain})
            with col_b:
                domain = st.selectbox("Domain", ["All"] + domains, key="segment_domain")
            with col_c:
                page_size = st.number_input(
                    "Page size",
                    min_value=20,
                    max_value=2000,
                    value=PAGE_SIZE_DEFAULT,
                    step=20,
                )

            filtered = filter_records(records, search=search, domain=domain)
            total_pages = max(1, (len(filtered) + int(page_size) - 1) // int(page_size))
            page = st.number_input(
                "Page",
                min_value=1,
                max_value=total_pages,
                value=min(st.session_state.current_page, total_pages),
                step=1,
            )
            st.session_state.current_page = int(page)

            start = (int(page) - 1) * int(page_size)
            end = start + int(page_size)
            page_records = filtered[start:end]

            edited_df = show_record_table(page_records, "segments_editor", editable=True)

            col1, col2 = st.columns([1, 2])
            with col1:
                if st.button("Save Table Edits", use_container_width=True):
                    apply_table_edits(edited_df, records, whole_word)
                    st.success("Edits saved.")
                    st.rerun()

            with col2:
                delete_id_text = st.text_input("Record ID(s) to delete", placeholder="e.g. 4, 8, 15")
                if st.button("Delete Selected Record ID(s)", use_container_width=True):
                    ids = {int(x.strip()) for x in delete_id_text.split(",") if x.strip().isdigit()}
                    if ids:
                        before = len(records)
                        st.session_state.records = [r for r in records if r.record_id not in ids]
                        st.session_state.stats = recalculate_all(st.session_state.records, whole_word)
                        removed = before - len(st.session_state.records)
                        log(f"Deleted {removed} selected record(s).")
                        st.success(f"Deleted {removed} record(s).")
                        st.rerun()
                    else:
                        st.warning("Enter at least one valid Record ID.")
        else:
            st.info("Upload TMX files and run the analysis first.")

    with tab_duplicates:
        st.subheader("Duplicates")
        if records:
            col_a, col_b = st.columns([2, 1])
            with col_a:
                search = st.text_input("Search duplicates", key="duplicate_search")
            with col_b:
                dtype = st.selectbox(
                    "Duplicate Type",
                    ["All", "Exact source+target duplicate", "Same source, different target", "Normalized source duplicate"],
                    key="duplicate_type",
                )
            filtered = filter_records(records, search=search, duplicate_type=dtype, only_duplicates=True)
            show_record_table(filtered, "duplicates_table")
        else:
            st.info("Upload TMX files and run the analysis first.")

    with tab_glossary:
        st.subheader("Glossary")
        if st.session_state.glossary_name:
            st.caption(f"Loaded glossary: {st.session_state.glossary_name}")

        if records:
            col_a, col_b = st.columns([2, 1])
            with col_a:
                search = st.text_input("Search glossary matches or violations", key="glossary_search")
            with col_b:
                gstatus = st.selectbox("Status", ["All", "Matched", "Violation", "No term hit"], key="glossary_status")
            filtered = filter_records(records, search=search, glossary_status=gstatus, only_glossary=True)
            show_record_table(filtered, "glossary_table")
        else:
            st.info("Upload TMX files and run the analysis first.")

    with tab_qa:
        st.subheader("QA")
        if records:
            col_a, col_b = st.columns([2, 1])
            with col_a:
                search = st.text_input("Search QA issues", key="qa_search")
            with col_b:
                qstatus = st.selectbox("Status", ["All", "Issues", "OK"], key="qa_status")
            filtered = filter_records(records, search=search, qa_status=qstatus, only_qa=(qstatus != "OK"))
            show_record_table(filtered, "qa_table")
        else:
            st.info("Upload TMX files and run the analysis first.")

    with tab_resolver:
        st.subheader("Duplicate Resolver")
        dup_records = [r for r in records if r.duplicate_group]

        if not dup_records:
            st.info("There are no duplicate groups to resolve.")
        else:
            groups: Dict[str, List[SegmentRecord]] = defaultdict(list)
            for r in dup_records:
                groups[r.duplicate_group].append(r)

            group_id = st.selectbox("Duplicate Group", sorted(groups.keys()))
            group_records = groups[group_id]
            show_record_table(group_records, "resolver_table")

            keep_id = st.selectbox("Record to keep", [r.record_id for r in group_records])

            if st.button("Keep Selected / Delete Others", type="primary"):
                delete_ids = {r.record_id for r in group_records if r.record_id != keep_id}
                st.session_state.records = [r for r in records if r.record_id not in delete_ids]
                st.session_state.stats = recalculate_all(st.session_state.records, whole_word)
                log(f"Duplicate resolver kept record {keep_id} and deleted {len(delete_ids)} record(s).")
                st.success(f"Kept record {keep_id}; deleted {len(delete_ids)} record(s).")
                st.rerun()

    with tab_logs:
        st.subheader("Logs")
        st.text_area("Application Log", "\n".join(st.session_state.logs), height=520)


if __name__ == "__main__":
    app()
